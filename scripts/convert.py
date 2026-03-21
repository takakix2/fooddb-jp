"""
convert.py — 日本食品標準成分表 Excel → クリーン JSONL 変換パイプライン
v2.0.0

Row 12 の INFOODS 識別子を直接キーとして使用し、
旧版のヘッダー結合ロジックを排除したクリーン版。

Usage:
    python convert.py [--excel PATH] [--output-dir PATH]
"""

import pandas as pd
import json
import re
import argparse
from pathlib import Path


# ========================================
# 定数
# ========================================

EXCEL_FILE = Path(__file__).parent.parent / "data" / "raw" / "成分表.xlsx"
OUTPUT_DIR = Path(__file__).parent.parent / "data" / "output"
SHEET_NAME = "表全体"

# Excel の行番号（0-indexed for pandas skiprows）
HEADER_ROW_INDEX = 11  # Row 12 (成分識別子)
DATA_START_ROW = 12    # Row 13 (データ開始)

# 固定列（A〜D, BJ）のマッピング
FIXED_COLUMNS = {
    0: "group",         # A: 食品群番号
    1: "food_number",   # B: 食品番号
    2: "index_number",  # C: 索引番号
    3: "food_name",     # D: 食品名
}
REMARKS_COL = 61  # BJ列（備考）

# INFOODS 識別子 → 日本語ラベル・単位のマスター定義
# Excel Row 3-4 (日本語見出し), Row 11 (単位), Row 12 (識別子) から構築
NUTRIENT_MASTER = {
    "REFUSE":     {"label_jp": "廃棄率",                     "label_en": "Refuse",                                "unit": "%"},
    "ENERC":      {"label_jp": "エネルギー",                   "label_en": "Energy",                                "unit": "kJ"},
    "ENERC_KCAL": {"label_jp": "エネルギー",                   "label_en": "Energy",                                "unit": "kcal"},
    "WATER":      {"label_jp": "水分",                       "label_en": "Water",                                 "unit": "g"},
    "PROTCAA":    {"label_jp": "アミノ酸組成によるたんぱく質",     "label_en": "Protein (amino acid composition)",      "unit": "g"},
    "PROT-":      {"label_jp": "たんぱく質",                   "label_en": "Protein",                               "unit": "g"},
    "FATNLEA":    {"label_jp": "脂肪酸のトリアシルグリセロール当量", "label_en": "Fat (NLEA, triacylglycerol equiv.)",    "unit": "g"},
    "CHOLE":      {"label_jp": "コレステロール",                "label_en": "Cholesterol",                           "unit": "mg"},
    "FAT-":       {"label_jp": "脂質",                       "label_en": "Total fat",                             "unit": "g"},
    "CHOAVLM":    {"label_jp": "利用可能炭水化物（単糖当量）",     "label_en": "Available carbohydrate (monosacch.)",   "unit": "g"},
    "CHOAVL":     {"label_jp": "利用可能炭水化物（質量計）",       "label_en": "Available carbohydrate (by weight)",    "unit": "g"},
    "CHOAVLDF-":  {"label_jp": "差引き法による利用可能炭水化物",    "label_en": "Available carbohydrate (by diff.)",     "unit": "g"},
    "FIB-":       {"label_jp": "食物繊維総量",                  "label_en": "Total dietary fibre",                   "unit": "g"},
    "POLYL":      {"label_jp": "糖アルコール",                  "label_en": "Polyols",                               "unit": "g"},
    "CHOCDF-":    {"label_jp": "炭水化物",                     "label_en": "Total carbohydrate (by diff.)",          "unit": "g"},
    "OA":         {"label_jp": "有機酸",                      "label_en": "Organic acids",                         "unit": "g"},
    "ASH":        {"label_jp": "灰分",                       "label_en": "Ash",                                   "unit": "g"},
    "NA":         {"label_jp": "ナトリウム",                   "label_en": "Sodium",                                "unit": "mg"},
    "K":          {"label_jp": "カリウム",                     "label_en": "Potassium",                             "unit": "mg"},
    "CA":         {"label_jp": "カルシウム",                    "label_en": "Calcium",                               "unit": "mg"},
    "MG":         {"label_jp": "マグネシウム",                  "label_en": "Magnesium",                             "unit": "mg"},
    "P":          {"label_jp": "リン",                        "label_en": "Phosphorus",                            "unit": "mg"},
    "FE":         {"label_jp": "鉄",                         "label_en": "Iron",                                  "unit": "mg"},
    "ZN":         {"label_jp": "亜鉛",                       "label_en": "Zinc",                                  "unit": "mg"},
    "CU":         {"label_jp": "銅",                         "label_en": "Copper",                                "unit": "mg"},
    "MN":         {"label_jp": "マンガン",                     "label_en": "Manganese",                             "unit": "mg"},
    "ID":         {"label_jp": "ヨウ素",                      "label_en": "Iodine",                                "unit": "µg"},
    "SE":         {"label_jp": "セレン",                      "label_en": "Selenium",                              "unit": "µg"},
    "CR":         {"label_jp": "クロム",                      "label_en": "Chromium",                              "unit": "µg"},
    "MO":         {"label_jp": "モリブデン",                   "label_en": "Molybdenum",                            "unit": "µg"},
    "RETOL":      {"label_jp": "レチノール",                   "label_en": "Retinol",                               "unit": "µg"},
    "CARTA":      {"label_jp": "α-カロテン",                  "label_en": "alpha-Carotene",                        "unit": "µg"},
    "CARTB":      {"label_jp": "β-カロテン",                  "label_en": "beta-Carotene",                         "unit": "µg"},
    "CRYPXB":     {"label_jp": "β-クリプトキサンチン",           "label_en": "beta-Cryptoxanthin",                    "unit": "µg"},
    "CARTBEQ":    {"label_jp": "β-カロテン当量",               "label_en": "beta-Carotene equivalent",              "unit": "µg"},
    "VITA_RAE":   {"label_jp": "レチノール活性当量",             "label_en": "Vitamin A (RAE)",                       "unit": "µg"},
    "VITD":       {"label_jp": "ビタミンD",                   "label_en": "Vitamin D",                             "unit": "µg"},
    "TOCPHA":     {"label_jp": "α-トコフェロール",              "label_en": "alpha-Tocopherol",                      "unit": "mg"},
    "TOCPHB":     {"label_jp": "β-トコフェロール",              "label_en": "beta-Tocopherol",                       "unit": "mg"},
    "TOCPHG":     {"label_jp": "γ-トコフェロール",              "label_en": "gamma-Tocopherol",                      "unit": "mg"},
    "TOCPHD":     {"label_jp": "δ-トコフェロール",              "label_en": "delta-Tocopherol",                      "unit": "mg"},
    "VITK":       {"label_jp": "ビタミンK",                   "label_en": "Vitamin K",                             "unit": "µg"},
    "THIA":       {"label_jp": "ビタミンB1",                  "label_en": "Thiamin",                               "unit": "mg"},
    "RIBF":       {"label_jp": "ビタミンB2",                  "label_en": "Riboflavin",                            "unit": "mg"},
    "NIA":        {"label_jp": "ナイアシン",                   "label_en": "Niacin",                                "unit": "mg"},
    "NE":         {"label_jp": "ナイアシン当量",                "label_en": "Niacin equivalent",                     "unit": "mg"},
    "VITB6A":     {"label_jp": "ビタミンB6",                  "label_en": "Vitamin B-6",                           "unit": "mg"},
    "VITB12":     {"label_jp": "ビタミンB12",                 "label_en": "Vitamin B-12",                          "unit": "µg"},
    "FOL":        {"label_jp": "葉酸",                       "label_en": "Folate",                                "unit": "µg"},
    "PANTAC":     {"label_jp": "パントテン酸",                 "label_en": "Pantothenic acid",                      "unit": "mg"},
    "BIOT":       {"label_jp": "ビオチン",                    "label_en": "Biotin",                                "unit": "µg"},
    "VITC":       {"label_jp": "ビタミンC",                   "label_en": "Vitamin C",                             "unit": "mg"},
    "ALC":        {"label_jp": "アルコール",                   "label_en": "Alcohol",                               "unit": "g"},
    "NACL_EQ":    {"label_jp": "食塩相当量",                   "label_en": "Salt equivalent",                       "unit": "g"},
}

# 食品群マスター
FOOD_GROUPS = {
    "01": "穀類",
    "02": "いも及びでん粉類",
    "03": "砂糖及び甘味類",
    "04": "豆類",
    "05": "種実類",
    "06": "野菜類",
    "07": "果実類",
    "08": "きのこ類",
    "09": "藻類",
    "10": "魚介類",
    "11": "肉類",
    "12": "卵類",
    "13": "乳類",
    "14": "油脂類",
    "15": "菓子類",
    "16": "し好飲料類",
    "17": "調味料及び香辛料類",
    "18": "調理済み流通食品類",
}


# ========================================
# 値クリーニング（旧コードから最良部分を継承）
# ========================================

def clean_value(val) -> float | str | None:
    """Excel セルの値をクリーニングする。

    処理ルール:
    - NaN, "-", 空文字 → None
    - "Tr", "(Tr)" → 0.0 (微量)
    - "(1.5)" のような括弧付き → 1.5 (推定値, フラグ別管理)
    - "*" → None (未測定)
    - 数値変換可能 → float
    - それ以外 → 文字列のまま
    """
    if pd.isna(val):
        return None
    val_str = str(val).strip()
    if val_str in ("-", "", "NaN", "*"):
        return None
    if val_str.lower() in ("tr", "(tr)"):
        return 0.0
    # 括弧付き推定値: (1.5) → 1.5
    is_estimated = False
    if val_str.startswith("(") and val_str.endswith(")"):
        val_str = val_str[1:-1].strip()
        is_estimated = True
    try:
        return float(val_str)
    except ValueError:
        return val_str


def is_estimated(val) -> bool:
    """値が推定値（括弧付き）かどうかを判定する。"""
    if pd.isna(val):
        return False
    val_str = str(val).strip()
    return val_str.startswith("(") and val_str.endswith(")")


# ========================================
# メイン変換パイプライン
# ========================================

def extract_column_mapping(excel_path: Path) -> dict[int, str]:
    """Row 12 から列インデックス → INFOODS 識別子のマッピングを構築する。"""
    import openpyxl
    wb = openpyxl.load_workbook(str(excel_path), read_only=True)
    ws = wb[SHEET_NAME]

    row12 = list(list(ws.iter_rows(
        min_row=12, max_row=12,
        max_col=ws.max_column,
        values_only=True
    ))[0])
    wb.close()

    mapping = {}
    for col_idx, cell_val in enumerate(row12):
        if cell_val is not None:
            tag = str(cell_val).strip()
            if tag and tag != "成分識別子":
                mapping[col_idx] = tag
    return mapping


def parse_food_name(raw_name: str) -> dict:
    """食品名を分類パーツに分解する。

    例: "＜穀類＞ ［こむぎ］ 薄力粉　1等" →
        {"subgroup": "穀類", "group": "こむぎ", "category": "薄力粉", "item": "1等"}
    """
    result = {"subgroup": None, "group": None, "category": None, "item": None}
    name = raw_name or ""

    # ＜サブグループ＞
    m = re.search(r"＜(.+?)＞", name)
    if m:
        result["subgroup"] = m.group(1)
        name = name[m.end():].strip()

    # ［グループ］
    m = re.search(r"［(.+?)］", name)
    if m:
        result["group"] = m.group(1)
        name = name[m.end():].strip()

    # カテゴリ　アイテム（全角スペース区切り）
    parts = name.split("　")
    if len(parts) >= 1:
        result["category"] = parts[0].strip() or None
    if len(parts) >= 2:
        result["item"] = "　".join(parts[1:]).strip() or None

    return result


def convert(excel_path: Path, output_dir: Path) -> None:
    """Excel → クリーン JSONL 変換のメインエントリポイント。"""
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"📖 Excel読み込み: {excel_path}")

    # Step 1: 列マッピング取得
    col_mapping = extract_column_mapping(excel_path)
    print(f"✅ 成分列マッピング: {len(col_mapping)} 列検出")

    # Step 2: 成分定義マスター出力
    nutrient_defs = []
    for sort_order, (col_idx, tag) in enumerate(sorted(col_mapping.items())):
        master = NUTRIENT_MASTER.get(tag, {})
        nutrient_defs.append({
            "tag": tag,
            "label_jp": master.get("label_jp", tag),
            "label_en": master.get("label_en", tag),
            "unit": master.get("unit", ""),
            "sort_order": sort_order,
            "excel_col": col_idx,
        })

    defs_path = output_dir / "nutrient_defs.json"
    with open(defs_path, "w", encoding="utf-8") as f:
        json.dump(nutrient_defs, f, ensure_ascii=False, indent=2)
    print(f"✅ 成分定義マスター: {defs_path} ({len(nutrient_defs)} 成分)")

    # Step 3: データ行読み込み & 変換
    df = pd.read_excel(str(excel_path), sheet_name=SHEET_NAME, header=None, skiprows=DATA_START_ROW)
    print(f"📊 データ行: {len(df)} 行")

    foods_path = output_dir / "foods.jsonl"
    nutrients_path = output_dir / "food_nutrients.jsonl"
    food_count = 0
    nutrient_count = 0

    with (
        open(foods_path, "w", encoding="utf-8") as f_foods,
        open(nutrients_path, "w", encoding="utf-8") as f_nutrients,
    ):
        for _, row in df.iterrows():
            # 食品番号がない行はスキップ
            food_number = str(row[1]).strip() if pd.notna(row[1]) else None
            if not food_number or food_number == "nan":
                continue

            group_code = str(int(row[0])).zfill(2) if pd.notna(row[0]) else None
            index_number = str(row[2]).strip() if pd.notna(row[2]) else None
            food_name = str(row[3]).strip() if pd.notna(row[3]) else None
            remarks = clean_value(row[REMARKS_COL]) if REMARKS_COL < len(row) else None

            # 食品名の分類パース
            name_parts = parse_food_name(food_name)

            # foods レコード
            food_record = {
                "food_number": food_number,
                "index_number": index_number,
                "food_name": food_name,
                "group_code": group_code,
                "group_name": FOOD_GROUPS.get(group_code, ""),
                "subgroup": name_parts["subgroup"],
                "category": name_parts["category"],
                "item": name_parts["item"],
                "remarks": remarks if isinstance(remarks, str) else None,
            }
            f_foods.write(json.dumps(food_record, ensure_ascii=False) + "\n")
            food_count += 1

            # nutrients レコード（成分ごとに1行）
            for col_idx, tag in col_mapping.items():
                if col_idx >= len(row):
                    continue
                raw_val = row[col_idx]
                value = clean_value(raw_val)
                if value is None:
                    continue
                if not isinstance(value, (int, float)):
                    continue  # 数値以外はスキップ（食品名列等）

                nutrient_record = {
                    "food_number": food_number,
                    "tag": tag,
                    "value": value,
                    "estimated": is_estimated(raw_val),
                }
                f_nutrients.write(json.dumps(nutrient_record, ensure_ascii=False) + "\n")
                nutrient_count += 1

    print(f"✅ 食品マスター: {foods_path} ({food_count} 件)")
    print(f"✅ 成分データ: {nutrients_path} ({nutrient_count} レコード)")

    # Step 4: サマリー出力
    summary = {
        "source": "日本食品標準成分表（八訂）増補2023年",
        "source_url": "https://www.mext.go.jp/a_menu/syokuhinseibun/mext_00001.html",
        "foods_count": food_count,
        "nutrients_count": nutrient_count,
        "nutrient_defs_count": len(nutrient_defs),
        "excel_file": str(excel_path.name),
        "version": "2.0.0",
    }
    summary_path = output_dir / "conversion_summary.json"
    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    print(f"\n📋 変換サマリー: {summary_path}")
    print(f"   食品数: {food_count}, 成分レコード数: {nutrient_count}")


# ========================================
# CLI
# ========================================

def main():
    parser = argparse.ArgumentParser(description="日本食品標準成分表 Excel → JSONL 変換")
    parser.add_argument("--excel", type=Path, default=EXCEL_FILE, help="入力 Excel ファイルパス")
    parser.add_argument("--output-dir", type=Path, default=OUTPUT_DIR, help="出力ディレクトリ")
    args = parser.parse_args()

    convert(args.excel, args.output_dir)


if __name__ == "__main__":
    main()
