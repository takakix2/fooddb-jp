"""
convert_all.py — 日本食品標準成分表 全テーブル統合変換
v1.0.0

本表 + アミノ酸(4) + 脂肪酸(3) + 炭水化物(3) の全11 Excelファイルを
統合してクリーンな JSONL + SQLite に変換する。

Usage:
    python convert_all.py [--raw-dir PATH] [--output-dir PATH]
"""

import openpyxl
import pandas as pd
import json
import re
import sqlite3
import argparse
from pathlib import Path

# ========================================
# 定数: 全テーブル定義
# ========================================

BASE_DIR = Path(__file__).parent.parent
RAW_DIR = BASE_DIR / "data" / "raw"
OUTPUT_DIR = BASE_DIR / "data" / "output"

# 各Excelファイルの構造定義
TABLE_DEFS = [
    {
        "file": "成分表.xlsx",
        "table_id": "main",
        "table_name": "本表（一般成分）",
        "table_name_en": "General Composition",
        "sheet": "表全体",
        "id_row": 12,       # 1-indexed: 成分識別子が入っている行
        "data_start": 13,   # 1-indexed: データ開始行
        "food_cols": {"group": 0, "food_number": 1, "index": 2, "name": 3},
        "remarks_col": 61,  # BJ列
    },
    {
        "file": "amino_acid_table1.xlsx",
        "table_id": "amino1",
        "table_name": "アミノ酸成分表 第1表（可食部100g当たり）",
        "table_name_en": "Amino Acids per 100g edible portion",
        "sheet": "表全体",
        "id_row": 5,
        "data_start": 7,
        "food_cols": {"group": 0, "food_number": 1, "index": 2, "name": 3},
        "remarks_col": None,
    },
    {
        "file": "amino_acid_table2.xlsx",
        "table_id": "amino2",
        "table_name": "アミノ酸成分表 第2表（基準窒素1g当たり）",
        "table_name_en": "Amino Acids per 1g reference nitrogen",
        "sheet": "表全体",
        "id_row": 5,
        "data_start": 7,
        "food_cols": {"group": 0, "food_number": 1, "index": 2, "name": 3},
        "remarks_col": None,
    },
    {
        "file": "amino_acid_table3.xlsx",
        "table_id": "amino3",
        "table_name": "アミノ酸成分表 第3表（たんぱく質1g当たり）",
        "table_name_en": "Amino Acids per 1g protein (amino acid composition)",
        "sheet": "表全体",
        "id_row": 5,
        "data_start": 7,
        "food_cols": {"group": 0, "food_number": 1, "index": 2, "name": 3},
        "remarks_col": None,
    },
    {
        "file": "amino_acid_table4.xlsx",
        "table_id": "amino4",
        "table_name": "アミノ酸成分表 第4表（たんぱく質1g当たり、基準窒素）",
        "table_name_en": "Amino Acids per 1g protein (reference nitrogen)",
        "sheet": "表全体",
        "id_row": 5,
        "data_start": 7,
        "food_cols": {"group": 0, "food_number": 1, "index": 2, "name": 3},
        "remarks_col": None,
    },
    {
        "file": "fatty_acid_table1.xlsx",
        "table_id": "fatty1",
        "table_name": "脂肪酸成分表 第1表（可食部100g当たり）",
        "table_name_en": "Fatty Acids per 100g edible portion",
        "sheet": "表全体",
        "id_row": 5,
        "data_start": 13,
        "food_cols": {"group": 0, "food_number": 1, "index": 2, "name": 3},
        "remarks_col": None,
    },
    {
        "file": "fatty_acid_table2.xlsx",
        "table_id": "fatty2",
        "table_name": "脂肪酸成分表 第2表（脂肪酸100g当たり）",
        "table_name_en": "Fatty Acids per 100g fatty acids",
        "sheet": "表全体",
        "id_row": 5,
        "data_start": 13,
        "food_cols": {"group": 0, "food_number": 1, "index": 2, "name": 3},
        "remarks_col": None,
    },
    {
        "file": "fatty_acid_table3.xlsx",
        "table_id": "fatty3",
        "table_name": "脂肪酸成分表 第3表（脂質1g当たり）",
        "table_name_en": "Fatty Acids per 1g lipid",
        "sheet": "表全体",
        "id_row": 5,
        "data_start": 13,
        "food_cols": {"group": 0, "food_number": 1, "index": 2, "name": 3},
        "remarks_col": None,
    },
    {
        "file": "carbohydrate_main.xlsx",
        "table_id": "carb_main",
        "table_name": "炭水化物成分表 本表（可食部100g当たり）",
        "table_name_en": "Carbohydrates per 100g edible portion",
        "sheet": "表全体",
        "id_row": 5,
        "data_start": 13,
        "food_cols": {"group": 0, "food_number": 1, "index": 2, "name": 3},
        "remarks_col": None,
    },
    {
        "file": "carbohydrate_appendix1.xlsx",
        "table_id": "carb_fiber",
        "table_name": "炭水化物成分表 別表1（食物繊維）",
        "table_name_en": "Dietary Fibre",
        "sheet": "表全体",
        "id_row": 8,
        "data_start": 13,
        "food_cols": {"group": 0, "food_number": 1, "index": 2, "name": 3},
        "remarks_col": None,
    },
    {
        "file": "carbohydrate_appendix2.xlsx",
        "table_id": "carb_organic",
        "table_name": "炭水化物成分表 別表2（有機酸）",
        "table_name_en": "Organic Acids",
        "sheet": "表全体",
        "id_row": 5,
        "data_start": 13,
        "food_cols": {"group": 0, "food_number": 1, "index": 2, "name": 3},
        "remarks_col": None,
    },
]

FOOD_GROUPS = {
    "01": "穀類", "02": "いも及びでん粉類", "03": "砂糖及び甘味類",
    "04": "豆類", "05": "種実類", "06": "野菜類", "07": "果実類",
    "08": "きのこ類", "09": "藻類", "10": "魚介類", "11": "肉類",
    "12": "卵類", "13": "乳類", "14": "油脂類", "15": "菓子類",
    "16": "し好飲料類", "17": "調味料及び香辛料類", "18": "調理済み流通食品類",
}


# ========================================
# ユーティリティ
# ========================================

def parse_food_name(name: str) -> dict:
    """文科省の食品標準成分表の食品名を構造化パースする。

    命名パターン:
      [＜category＞] base_name [［subcategory］] [detail...]
    """
    remaining = name.strip()
    category = subcategory = base_name = detail = None

    m = re.match(r'^＜([^＞]+)＞[\s\u3000]*(.*)', remaining)
    if m:
        category = m.group(1)
        remaining = m.group(2)

    m_sub = re.search(r'［([^\]]+)］', remaining)
    if m_sub:
        subcategory = m_sub.group(1)
        before = remaining[:m_sub.start()].strip().rstrip('\u3000').rstrip()
        after = remaining[m_sub.end():].strip().lstrip('\u3000').lstrip()
        base_name = before if before else None
        detail = after if after else None
    else:
        tokens = re.split(r'[\s\u3000]+', remaining, maxsplit=1)
        if tokens:
            base_name = tokens[0] if tokens[0] else None
            if len(tokens) > 1 and tokens[1]:
                detail = tokens[1]

    return {"category": category, "subcategory": subcategory,
            "base_name": base_name, "detail": detail}

def clean_value(val) -> float | None:
    """Excel セルの値を数値に変換。数値以外は None。"""
    if pd.isna(val):
        return None
    val_str = str(val).strip()
    if val_str in ("-", "", "NaN", "*", "…"):
        return None
    if val_str.lower() in ("tr", "(tr)"):
        return 0.0
    if val_str.startswith("(") and val_str.endswith(")"):
        val_str = val_str[1:-1].strip()
    try:
        return float(val_str)
    except ValueError:
        return None


def is_estimated(val) -> bool:
    if pd.isna(val):
        return False
    val_str = str(val).strip()
    return val_str.startswith("(") and val_str.endswith(")")


def extract_id_row(excel_path: Path, sheet: str, id_row: int, max_col: int) -> dict[int, str]:
    """指定行から列インデックス → タグ名のマッピングを取得。"""
    wb = openpyxl.load_workbook(str(excel_path), read_only=True)
    ws = wb[sheet]
    row_data = list(list(ws.iter_rows(
        min_row=id_row, max_row=id_row,
        max_col=max_col, values_only=True
    ))[0])
    wb.close()

    mapping = {}
    skip_labels = {"成分識別子", "単位", None}
    for col_idx, cell in enumerate(row_data):
        if cell is not None:
            tag = str(cell).strip()
            if tag and tag not in skip_labels:
                mapping[col_idx] = tag
    return mapping


def extract_unit_row(excel_path: Path, sheet: str, unit_row: int, max_col: int) -> dict[int, str]:
    """単位行から列インデックス → 単位のマッピングを取得。"""
    wb = openpyxl.load_workbook(str(excel_path), read_only=True)
    ws = wb[sheet]
    row_data = list(list(ws.iter_rows(
        min_row=unit_row, max_row=unit_row,
        max_col=max_col, values_only=True
    ))[0])
    wb.close()

    mapping = {}
    for col_idx, cell in enumerate(row_data):
        if cell is not None:
            unit = str(cell).strip()
            # 装飾文字除去
            import re
            unit = re.sub(r'[…(）() 　]', '', unit)
            if unit and unit not in ("成分識別子", "単位"):
                mapping[col_idx] = unit
    return mapping


# ========================================
# メイン処理
# ========================================

def process_all(raw_dir: Path, output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)

    all_nutrient_defs = {}  # tag -> {table_id, label_jp, unit, ...}
    all_nutrients = []       # [{food_number, tag, value, estimated, table_id}, ...]
    all_foods = {}           # food_number -> {food_name, group_code, ...}

    for tdef in TABLE_DEFS:
        excel_path = raw_dir / tdef["file"]
        if not excel_path.exists():
            print(f"⚠️  スキップ: {tdef['file']} (ファイルなし)")
            continue

        print(f"\n{'='*50}")
        print(f"📄 {tdef['table_name']}")
        print(f"   ファイル: {tdef['file']}")

        # 列マッピング取得
        df = pd.read_excel(str(excel_path), sheet_name=tdef["sheet"], header=None,
                           skiprows=tdef["data_start"] - 1)
        max_col = df.shape[1]

        col_mapping = extract_id_row(excel_path, tdef["sheet"], tdef["id_row"], max_col)

        # 食品列を除外
        food_col_indices = set(tdef["food_cols"].values())
        if tdef.get("remarks_col") is not None:
            food_col_indices.add(tdef["remarks_col"])

        nutrient_cols = {
            col: tag for col, tag in col_mapping.items()
            if col not in food_col_indices
        }

        print(f"   成分列: {len(nutrient_cols)} 列, データ行: {len(df)} 行")

        # 成分定義を収集
        for col_idx, tag in nutrient_cols.items():
            if tag not in all_nutrient_defs:
                all_nutrient_defs[tag] = {
                    "tag": tag,
                    "table_id": tdef["table_id"],
                    "table_name": tdef["table_name"],
                }

        # データ行処理
        food_count = 0
        nutrient_count = 0

        for _, row in df.iterrows():
            fc = tdef["food_cols"]
            food_number = str(row[fc["food_number"]]).strip() if pd.notna(row[fc["food_number"]]) else None
            if not food_number or food_number == "nan":
                continue

            # 食品マスターに追加（本表から優先）
            if food_number not in all_foods:
                group_code = str(int(row[fc["group"]])).zfill(2) if pd.notna(row[fc["group"]]) else None
                food_name = str(row[fc["name"]]).strip() if pd.notna(row[fc["name"]]) else None
                all_foods[food_number] = {
                    "food_number": food_number,
                    "index_number": str(row[fc["index"]]).strip() if pd.notna(row[fc["index"]]) else None,
                    "food_name": food_name,
                    "group_code": group_code,
                    "group_name": FOOD_GROUPS.get(group_code, ""),
                }

            # 成分値
            for col_idx, tag in nutrient_cols.items():
                if col_idx >= len(row):
                    continue
                raw_val = row[col_idx]
                value = clean_value(raw_val)
                if value is None:
                    continue
                all_nutrients.append({
                    "food_number": food_number,
                    "tag": tag,
                    "value": value,
                    "estimated": is_estimated(raw_val),
                    "table_id": tdef["table_id"],
                })
                nutrient_count += 1

            food_count += 1

        print(f"   ✅ 食品: {food_count} 件, 成分レコード: {nutrient_count}")

    # ========================================
    # 出力
    # ========================================

    print(f"\n{'='*50}")
    print(f"📊 統合結果")
    print(f"   食品マスター: {len(all_foods)} 件")
    print(f"   成分定義: {len(all_nutrient_defs)} タグ")
    print(f"   成分レコード: {len(all_nutrients)} 件")

    # JSONL出力
    foods_path = output_dir / "foods_all.jsonl"
    with open(foods_path, "w", encoding="utf-8") as f:
        for food in sorted(all_foods.values(), key=lambda x: x["food_number"]):
            f.write(json.dumps(food, ensure_ascii=False) + "\n")
    print(f"   → {foods_path.name}")

    nutrients_path = output_dir / "nutrients_all.jsonl"
    with open(nutrients_path, "w", encoding="utf-8") as f:
        for n in all_nutrients:
            f.write(json.dumps(n, ensure_ascii=False) + "\n")
    print(f"   → {nutrients_path.name}")

    defs_path = output_dir / "nutrient_defs_all.json"
    with open(defs_path, "w", encoding="utf-8") as f:
        json.dump(list(all_nutrient_defs.values()), f, ensure_ascii=False, indent=2)
    print(f"   → {defs_path.name}")

    # ========================================
    # SQLite出力
    # ========================================

    db_path = BASE_DIR / "fooddb.sqlite"
    print(f"\n💾 SQLite: {db_path}")

    conn = sqlite3.connect(str(db_path))
    cur = conn.cursor()

    # テーブル作成
    cur.executescript("""
        DROP TABLE IF EXISTS food_nutrients;
        DROP TABLE IF EXISTS nutrient_defs;
        DROP TABLE IF EXISTS foods;
        DROP TABLE IF EXISTS tables;

        CREATE TABLE tables (
            table_id   TEXT PRIMARY KEY,
            table_name TEXT NOT NULL
        );

        CREATE TABLE foods (
            food_number  TEXT PRIMARY KEY,
            index_number TEXT,
            food_name    TEXT NOT NULL,
            group_code   TEXT,
            group_name   TEXT,
            category     TEXT,
            subcategory  TEXT,
            base_name    TEXT,
            detail       TEXT
        );

        CREATE TABLE nutrient_defs (
            tag        TEXT PRIMARY KEY,
            table_id   TEXT NOT NULL,
            table_name TEXT,
            FOREIGN KEY (table_id) REFERENCES tables(table_id)
        );

        CREATE TABLE food_nutrients (
            food_number TEXT NOT NULL,
            tag         TEXT NOT NULL,
            value       REAL NOT NULL,
            estimated   INTEGER NOT NULL DEFAULT 0,
            table_id    TEXT NOT NULL,
            PRIMARY KEY (food_number, tag, table_id),
            FOREIGN KEY (food_number) REFERENCES foods(food_number),
            FOREIGN KEY (tag) REFERENCES nutrient_defs(tag)
        );
    """)

    # テーブル定義
    for tdef in TABLE_DEFS:
        cur.execute("INSERT OR IGNORE INTO tables VALUES (?, ?)",
                    (tdef["table_id"], tdef["table_name"]))

    # 食品
    for food in all_foods.values():
        parsed = parse_food_name(food["food_name"]) if food["food_name"] else {}
        cur.execute("INSERT OR IGNORE INTO foods VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (food["food_number"], food["index_number"],
                     food["food_name"], food["group_code"], food["group_name"],
                     parsed.get("category"), parsed.get("subcategory"),
                     parsed.get("base_name"), parsed.get("detail")))

    # 成分定義
    for d in all_nutrient_defs.values():
        cur.execute("INSERT OR IGNORE INTO nutrient_defs VALUES (?, ?, ?)",
                    (d["tag"], d["table_id"], d["table_name"]))

    # 成分値（バッチ挿入）
    cur.executemany(
        "INSERT OR IGNORE INTO food_nutrients VALUES (?, ?, ?, ?, ?)",
        [(n["food_number"], n["tag"], n["value"],
          1 if n["estimated"] else 0, n["table_id"])
         for n in all_nutrients]
    )

    conn.commit()

    # 統計クエリ
    cur.execute("SELECT COUNT(*) FROM foods")
    print(f"   foods: {cur.fetchone()[0]} 行")
    cur.execute("SELECT COUNT(*) FROM nutrient_defs")
    print(f"   nutrient_defs: {cur.fetchone()[0]} 行")
    cur.execute("SELECT COUNT(*) FROM food_nutrients")
    print(f"   food_nutrients: {cur.fetchone()[0]} 行")
    cur.execute("SELECT table_id, COUNT(*) FROM food_nutrients GROUP BY table_id ORDER BY table_id")
    print(f"\n   テーブル別レコード数:")
    for row in cur.fetchall():
        print(f"     {row[0]:15s}: {row[1]:>8,}")

    # インデックス追加
    cur.executescript("""
        CREATE INDEX IF NOT EXISTS idx_fn_food ON food_nutrients(food_number);
        CREATE INDEX IF NOT EXISTS idx_fn_tag ON food_nutrients(tag);
        CREATE INDEX IF NOT EXISTS idx_fn_table ON food_nutrients(table_id);
        CREATE INDEX IF NOT EXISTS idx_foods_group ON foods(group_code);
        CREATE INDEX IF NOT EXISTS idx_foods_category ON foods(category);
        CREATE INDEX IF NOT EXISTS idx_foods_subcategory ON foods(subcategory);
        CREATE INDEX IF NOT EXISTS idx_foods_base_name ON foods(base_name);
    """)
    conn.commit()

    db_size = db_path.stat().st_size
    print(f"\n   DBサイズ: {db_size/1024/1024:.1f} MB")
    conn.close()
    print("\n🎉 全テーブル統合完了！")


def main():
    parser = argparse.ArgumentParser(description="日本食品標準成分表 全テーブル統合変換")
    parser.add_argument("--raw-dir", type=Path, default=RAW_DIR)
    parser.add_argument("--output-dir", type=Path, default=OUTPUT_DIR)
    args = parser.parse_args()
    process_all(args.raw_dir, args.output_dir)


if __name__ == "__main__":
    main()
